from fastapi import FastAPI, HTTPException, Request, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, Response
from pydantic import BaseModel
from typing import List, Optional, Dict, Any, Union
import asyncio
import json
import os
import logging
from dotenv import load_dotenv
import numpy as np
from scipy import stats
from scipy.fft import fft
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import LineChart, BarChart, ScatterChart, Reference, Series
from openpyxl.utils import get_column_letter
from datetime import datetime
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend
import matplotlib.pyplot as plt
import subprocess
import tempfile
import shutil
import os

from llm_client import LLMClient
from stats_analyzer import StatsAnalyzer
from latex_pdf_generator import generate_latex_pdf

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

load_dotenv()
logger.info("Environment variables loaded")

# Get dummy data filename from environment variable, default to dummy_data.json
DUMMY_DATA_FILENAME = os.getenv("DUMMY_DATA_FILENAME", "dummy_data.json")
logger.info(f"Dummy data file: {DUMMY_DATA_FILENAME}")

# Log which API keys are available (without exposing the keys)
env_vars = {
    "OPENAI_API_KEY": "✓" if os.getenv("OPENAI_API_KEY") else "✗",
    "ANTHROPIC_API_KEY": "✓" if os.getenv("ANTHROPIC_API_KEY") else "✗",
    "DEEPSEEK_API_KEY": "✓" if os.getenv("DEEPSEEK_API_KEY") else "✗"
}
logger.info(f"API Keys status: {env_vars}")

app = FastAPI(title="LLM Random Number Generator")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "http://localhost:5173"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

llm_client = LLMClient()
stats_analyzer = StatsAnalyzer()


class PromptRequest(BaseModel):
    provider: str  # "openai", "anthropic", "deepseek"
    system_prompt: Optional[str] = None
    count: Optional[int] = 100
    api_key: Optional[str] = None  # API key for the selected provider
    batch_mode: Optional[bool] = False  # True = one request for all numbers, False = one request per number


class NumberData(BaseModel):
    numbers: List[float]
    provider: str


class MultiRunData(BaseModel):
    runs: List[List[float]]
    provider: str
    num_runs: int


@app.get("/")
async def root():
    return {"message": "LLM Random Number Generator API"}


@app.post("/generate")
async def generate_numbers(request: PromptRequest):
    """Generate random numbers from specified LLM provider"""
    logger.info(f"Received generate request: provider={request.provider}, count={request.count or 100}, has_api_key={bool(request.api_key)}, has_system_prompt={bool(request.system_prompt)}")
    if request.system_prompt:
        logger.info(f"System prompt received from frontend (full text): {request.system_prompt}")
        logger.debug(f"System prompt length: {len(request.system_prompt)} characters")
    else:
        logger.info("No system prompt provided, will use default prompt")
    try:
        numbers = await llm_client.generate_random_numbers(
            provider=request.provider,
            system_prompt=request.system_prompt,
            count=request.count or 100,
            api_key=request.api_key
        )
        logger.info(f"Generated {len(numbers)} numbers for provider {request.provider}")
        logger.debug(f"Numbers: {numbers[:10]}..." if len(numbers) > 10 else f"Numbers: {numbers}")
        return {"numbers": numbers, "provider": request.provider}
    except Exception as e:
        logger.error(f"Error generating numbers: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/generate/stream")
async def generate_numbers_stream(request: PromptRequest):
    """Stream random numbers from specified LLM provider"""
    logger.info(f"Received stream request: provider={request.provider}, count={request.count or 100}, batch_mode={request.batch_mode}, has_api_key={bool(request.api_key)}, has_system_prompt={bool(request.system_prompt)}")
    if request.system_prompt:
        logger.info(f"System prompt received from frontend (full text): {request.system_prompt}")
        logger.info(f"System prompt length: {len(request.system_prompt)} characters")
    else:
        logger.info("No system prompt provided, will use default prompt")

    async def event_generator():
        numbers_generated = 0
        try:
            async for number in llm_client.generate_random_numbers_stream(
                provider=request.provider,
                system_prompt=request.system_prompt,
                count=request.count or 100,
                api_key=request.api_key,
                batch_mode=request.batch_mode or False
            ):
                numbers_generated += 1
                logger.debug(f"Streaming number {numbers_generated}: {number}")
                yield f"data: {json.dumps({'number': number, 'provider': request.provider})}\n\n"
            logger.info(f"Stream completed: generated {numbers_generated} numbers")
        except Exception as e:
            logger.error(f"Error in stream: {str(e)}", exc_info=True)
            yield f"data: {json.dumps({'error': str(e)})}\n\n"
        finally:
            yield "data: [DONE]\n\n"
    
    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        }
    )


@app.post("/analyze")
async def analyze_numbers(request: Request):
    """Perform comprehensive statistical analysis on generated numbers"""
    try:
        body = await request.json()
        logger.info(f"Received analyze request. Keys in body: {list(body.keys())}")
        
        # Check if it's a multi-run request
        if 'runs' in body:
            runs_data = body.get('runs', [])
            logger.info(f"Multi-run analysis requested: {len(runs_data)} runs, num_runs={body.get('num_runs')}")
            if not runs_data:
                raise HTTPException(status_code=400, detail="'runs' array is empty")
            if not isinstance(runs_data, list):
                raise HTTPException(status_code=400, detail="'runs' must be an array")
            
            try:
                data = MultiRunData(**body)
                logger.info(f"Validated multi-run data: {len(data.runs)} runs, provider={data.provider}")
                # Log run lengths for debugging
                run_lengths = [len(run) for run in data.runs]
                logger.info(f"Run lengths: {run_lengths}")
                analysis = stats_analyzer.analyze_multi_run(data.runs, data.provider, data.num_runs)
                logger.info("Multi-run analysis completed successfully")
                return analysis
            except ValueError as e:
                logger.error(f"Validation error in multi-run analysis: {str(e)}", exc_info=True)
                raise HTTPException(status_code=400, detail=f"Validation error: {str(e)}")
            except Exception as e:
                logger.error(f"Error in multi-run analysis: {str(e)}", exc_info=True)
                raise HTTPException(status_code=500, detail=f"Multi-run analysis error: {str(e)}")
        elif 'numbers' in body:
            # Single run analysis (backward compatibility)
            numbers_data = body.get('numbers', [])
            logger.info(f"Single-run analysis requested: {len(numbers_data)} numbers")
            if not numbers_data:
                raise HTTPException(status_code=400, detail="'numbers' array is empty")
            if not isinstance(numbers_data, list):
                raise HTTPException(status_code=400, detail="'numbers' must be an array")
            
            try:
                data = NumberData(**body)
                logger.info(f"Validated single-run data: {len(data.numbers)} numbers, provider={data.provider}")
                analysis = stats_analyzer.analyze(data.numbers, data.provider)
                logger.info("Single-run analysis completed successfully")
                return analysis
            except ValueError as e:
                logger.error(f"Validation error in single-run analysis: {str(e)}", exc_info=True)
                raise HTTPException(status_code=400, detail=f"Validation error: {str(e)}")
            except Exception as e:
                logger.error(f"Error in single-run analysis: {str(e)}", exc_info=True)
                raise HTTPException(status_code=500, detail=f"Single-run analysis error: {str(e)}")
        else:
            logger.error(f"Invalid request body: missing 'runs' or 'numbers' key. Body keys: {list(body.keys())}")
            raise HTTPException(status_code=400, detail="Either 'runs' or 'numbers' must be provided")
    except HTTPException:
        # Re-raise HTTP exceptions as-is
        raise
    except json.JSONDecodeError as e:
        logger.error(f"Invalid JSON in request: {str(e)}")
        raise HTTPException(status_code=400, detail=f"Invalid JSON: {str(e)}")
    except Exception as e:
        logger.error(f"Unexpected error in analyze endpoint: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Analysis error: {str(e)}")


@app.get("/download/csv")
async def download_csv(
    runs: str = Query(..., description="JSON string of runs array"),
    provider: str = Query("manual", description="Provider name")
):
    """Download raw numbers as CSV file with columns run 1, run 2, etc."""
    try:
        import json
        runs_data = json.loads(runs)
        
        if not isinstance(runs_data, list) or len(runs_data) == 0:
            raise HTTPException(status_code=400, detail="Invalid runs data")
        
        # Create DataFrame with runs as columns
        max_len = max(len(run) for run in runs_data)
        data_dict = {}
        
        for i, run in enumerate(runs_data):
            col_name = f"run {i + 1}"
            # Pad shorter runs with empty strings
            padded_run = run + [""] * (max_len - len(run))
            data_dict[col_name] = padded_run
        
        df = pd.DataFrame(data_dict)
        
        # Create CSV in memory
        output = BytesIO()
        df.to_csv(output, index=False)
        output.seek(0)
        
        filename = f"random_numbers_{provider}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        return Response(
            content=output.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Invalid JSON in runs parameter")
    except Exception as e:
        logger.error(f"Error generating CSV: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error generating CSV: {str(e)}")


class XLSXDownloadRequest(BaseModel):
    analysis: Dict[str, Any]
    runs: List[List[float]]

@app.post("/download/xlsx")
async def download_xlsx(request: XLSXDownloadRequest):
    """Download full analysis report as XLSX file with metrics, charts data, and everything"""
    try:
        analysis = request.analysis
        runs_data = request.runs
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Helper function to create styled header
        def create_header(ws, row, col, text, fill_color="366092"):
            cell = ws.cell(row=row, column=col, value=text)
            cell.font = Font(bold=True, color="FFFFFF", size=12)
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            return cell
        
        # Helper function to write metric row
        def write_metric(ws, row, label, value, col=1):
            ws.cell(row=row, column=col, value=label).font = Font(bold=True)
            ws.cell(row=row, column=col + 1, value=value)
        
        # Sheet 1: Summary
        ws_summary = wb.create_sheet("Summary", 0)
        row = 1
        create_header(ws_summary, row, 1, "Analysis Summary", "366092")
        ws_summary.merge_cells(f"A{row}:B{row}")
        row += 2
        
        write_metric(ws_summary, row, "Provider:", analysis.get("provider", "N/A"))
        row += 1
        if "num_runs" in analysis:
            write_metric(ws_summary, row, "Number of Runs:", analysis.get("num_runs", "N/A"))
            row += 1
            write_metric(ws_summary, row, "Count per Run:", analysis.get("count_per_run", "N/A"))
            row += 1
        else:
            write_metric(ws_summary, row, "Count:", analysis.get("count", "N/A"))
            row += 1
        
        # Sheet 2: Raw Data
        ws_data = wb.create_sheet("Raw Data", 1)
        max_len = max(len(run) for run in runs_data) if runs_data else 0
        for i, run in enumerate(runs_data):
            col = i + 1
            header = ws_data.cell(row=1, column=col, value=f"run {i + 1}")
            header.font = Font(bold=True)
            header.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            for j, value in enumerate(run):
                ws_data.cell(row=j + 2, column=col, value=value)
        
        # Sheet 3: Basic Statistics
        ws_basic = wb.create_sheet("Basic Statistics", 2)
        row = 1
        create_header(ws_basic, row, 1, "Basic Descriptive Statistics", "70AD47")
        ws_basic.merge_cells(f"A{row}:B{row}")
        row += 2
        
        if "basic_stats" in analysis:
            stats = analysis["basic_stats"]
            metrics = [
                ("Mean", stats.get("mean")),
                ("Median", stats.get("median")),
                ("Std Dev", stats.get("std")),
                ("Variance", stats.get("variance")),
                ("Min", stats.get("min")),
                ("Max", stats.get("max")),
                ("Q25", stats.get("q25")),
                ("Q50", stats.get("q50")),
                ("Q75", stats.get("q75")),
                ("Q95", stats.get("q95")),
                ("Skewness", stats.get("skewness")),
                ("Kurtosis", stats.get("kurtosis"))
            ]
            for label, value in metrics:
                if value is not None:
                    write_metric(ws_basic, row, f"{label}:", f"{value:.6f}")
                    row += 1
        
        # Sheet 4: Distribution Analysis
        if "distribution" in analysis:
            ws_dist = wb.create_sheet("Distribution", 3)
            row = 1
            create_header(ws_dist, row, 1, "Distribution Analysis", "FFC000")
            ws_dist.merge_cells(f"A{row}:B{row}")
            row += 2
            
            dist = analysis["distribution"]
            
            # Uniformity test
            if "is_uniform" in dist:
                ws_dist.cell(row=row, column=1, value="Uniformity Test").font = Font(bold=True, size=11)
                row += 1
                write_metric(ws_dist, row, "Kolmogorov-Smirnov p-value:", f"{dist['is_uniform']['ks_p']:.6f}")
                row += 1
                row += 1
            
            # Histogram data and chart
            if "histogram" in dist:
                ws_dist.cell(row=row, column=1, value="Histogram Data").font = Font(bold=True, size=11)
                row += 1
                ws_dist.cell(row=row, column=1, value="Bin Edge").font = Font(bold=True)
                ws_dist.cell(row=row, column=2, value="Count").font = Font(bold=True)
                row += 1
                edges = dist["histogram"].get("edges", [])
                counts = dist["histogram"].get("counts", [])
                hist_start_row = row
                for i in range(len(counts)):
                    ws_dist.cell(row=row, column=1, value=f"{edges[i]:.6f}")
                    ws_dist.cell(row=row, column=2, value=counts[i])
                    row += 1
                
                # Add Histogram chart
                chart_hist = BarChart()
                chart_hist.type = "col"
                chart_hist.style = 10
                chart_hist.title = "Histogram"
                chart_hist.y_axis.title = "Count"
                chart_hist.x_axis.title = "Bin Edge"
                data = Reference(ws_dist, min_col=2, min_row=hist_start_row-1, max_row=row-1)
                cats = Reference(ws_dist, min_col=1, min_row=hist_start_row, max_row=row-1)
                chart_hist.add_data(data, titles_from_data=True)
                chart_hist.set_categories(cats)
                chart_hist.width = 15
                chart_hist.height = 10
                ws_dist.add_chart(chart_hist, f"D{hist_start_row}")
                row += 15
            
            # KDE data and chart
            if "kde" in dist:
                row += 2
                ws_dist.cell(row=row, column=1, value="Kernel Density Estimate (KDE)").font = Font(bold=True, size=11)
                row += 1
                ws_dist.cell(row=row, column=1, value="X").font = Font(bold=True)
                ws_dist.cell(row=row, column=2, value="Density").font = Font(bold=True)
                row += 1
                kde_x = dist["kde"].get("x", [])
                kde_y = dist["kde"].get("y", [])
                kde_start_row = row
                for x, y in zip(kde_x, kde_y):
                    ws_dist.cell(row=row, column=1, value=f"{x:.6f}")
                    ws_dist.cell(row=row, column=2, value=f"{y:.6f}")
                    row += 1
                
                # Add KDE chart
                chart_kde = LineChart()
                chart_kde.title = "Kernel Density Estimate"
                chart_kde.y_axis.title = "Density"
                chart_kde.x_axis.title = "Value"
                data = Reference(ws_dist, min_col=2, min_row=kde_start_row-1, max_row=row-1)
                cats = Reference(ws_dist, min_col=1, min_row=kde_start_row, max_row=row-1)
                chart_kde.add_data(data, titles_from_data=True)
                chart_kde.set_categories(cats)
                chart_kde.width = 15
                chart_kde.height = 10
                ws_dist.add_chart(chart_kde, f"D{kde_start_row}")
                row += 15
            
            # Q-Q plot data and chart
            if "qq_plot" in dist:
                row += 2
                ws_dist.cell(row=row, column=1, value="Q-Q Plot (Uniform Distribution)").font = Font(bold=True, size=11)
                row += 1
                ws_dist.cell(row=row, column=1, value="Theoretical").font = Font(bold=True)
                ws_dist.cell(row=row, column=2, value="Sample").font = Font(bold=True)
                row += 1
                theoretical = dist["qq_plot"].get("theoretical", [])
                sample = dist["qq_plot"].get("sample", [])
                qq_start_row = row
                for th, samp in zip(theoretical, sample):
                    ws_dist.cell(row=row, column=1, value=f"{th:.6f}")
                    ws_dist.cell(row=row, column=2, value=f"{samp:.6f}")
                    row += 1
                
                # Add Q-Q plot chart (using ScatterChart)
                chart_qq = ScatterChart()
                chart_qq.title = "Q-Q Plot (Uniform Distribution)"
                chart_qq.y_axis.title = "Sample"
                chart_qq.x_axis.title = "Theoretical"
                chart_qq.style = 13
                xvalues = Reference(ws_dist, min_col=1, min_row=qq_start_row, max_row=row-1)
                yvalues = Reference(ws_dist, min_col=2, min_row=qq_start_row-1, max_row=row-1)
                series = Series(yvalues, xvalues, title="Q-Q")
                chart_qq.series.append(series)
                chart_qq.width = 15
                chart_qq.height = 10
                ws_dist.add_chart(chart_qq, f"D{qq_start_row}")
        
        # Sheet 5: Range & Boundary Behavior
        if "range_behavior" in analysis:
            ws_range = wb.create_sheet("Range Behavior", 4)
            row = 1
            create_header(ws_range, row, 1, "Range & Boundary Behavior", "00B050")
            ws_range.merge_cells(f"A{row}:B{row}")
            row += 2
            
            range_beh = analysis["range_behavior"]
            
            # Boundary statistics
            if "boundaries" in range_beh:
                ws_range.cell(row=row, column=1, value="Boundary Statistics").font = Font(bold=True, size=11)
                row += 1
                boundaries = range_beh["boundaries"]
                write_metric(ws_range, row, "Min:", f"{boundaries.get('min', 0):.6f}")
                row += 1
                write_metric(ws_range, row, "Max:", f"{boundaries.get('max', 0):.6f}")
                row += 1
                write_metric(ws_range, row, "Near Min Count:", boundaries.get('near_min_count', 0))
                row += 1
                write_metric(ws_range, row, "Near Min %:", f"{boundaries.get('near_min_pct', 0):.6f}")
                row += 1
                write_metric(ws_range, row, "Near Max Count:", boundaries.get('near_max_count', 0))
                row += 1
                write_metric(ws_range, row, "Near Max %:", f"{boundaries.get('near_max_pct', 0):.6f}")
                row += 2
            
            # ECDF data and chart
            if "ecdf" in range_beh:
                ws_range.cell(row=row, column=1, value="Empirical Cumulative Distribution Function (ECDF)").font = Font(bold=True, size=11)
                row += 1
                ws_range.cell(row=row, column=1, value="X").font = Font(bold=True)
                ws_range.cell(row=row, column=2, value="ECDF").font = Font(bold=True)
                row += 1
                ecdf_x = range_beh["ecdf"].get("x", [])
                ecdf_y = range_beh["ecdf"].get("y", [])
                ecdf_start_row = row
                # Limit to reasonable number of points for Excel
                step = max(1, len(ecdf_x) // 1000) if len(ecdf_x) > 1000 else 1
                for i in range(0, len(ecdf_x), step):
                    ws_range.cell(row=row, column=1, value=f"{ecdf_x[i]:.6f}")
                    ws_range.cell(row=row, column=2, value=f"{ecdf_y[i]:.6f}")
                    row += 1
                
                # Add ECDF chart
                chart_ecdf = LineChart()
                chart_ecdf.title = "Empirical Cumulative Distribution Function"
                chart_ecdf.y_axis.title = "Cumulative Probability"
                chart_ecdf.x_axis.title = "Value"
                data = Reference(ws_range, min_col=2, min_row=ecdf_start_row-1, max_row=row-1)
                cats = Reference(ws_range, min_col=1, min_row=ecdf_start_row, max_row=row-1)
                chart_ecdf.add_data(data, titles_from_data=True)
                chart_ecdf.set_categories(cats)
                chart_ecdf.width = 15
                chart_ecdf.height = 10
                ws_range.add_chart(chart_ecdf, f"D{ecdf_start_row}")
        
        # Sheet 6: Aggregate Statistics (for multi-run)
        if "aggregate_stats" in analysis:
            ws_agg = wb.create_sheet("Aggregate Statistics", 5)
            row = 1
            create_header(ws_agg, row, 1, "Aggregate Statistics Across Runs", "7030A0")
            ws_agg.merge_cells(f"A{row}:D{row}")
            row += 1
            
            # Header row
            headers = ["Metric", "Mean Across Runs", "Std Dev Across Runs", "Range"]
            for col, header in enumerate(headers, 1):
                create_header(ws_agg, row, col, header, "B1A0C7")
            row += 1
            
            agg_stats = analysis["aggregate_stats"]
            metrics = ["mean", "std_dev", "skewness", "kurtosis"]
            for metric in metrics:
                if metric in agg_stats:
                    ws_agg.cell(row=row, column=1, value=metric.capitalize().replace("_", " ")).font = Font(bold=True)
                    ws_agg.cell(row=row, column=2, value=f"{agg_stats[metric].get('mean', 0):.6f}")
                    ws_agg.cell(row=row, column=3, value=f"{agg_stats[metric].get('std_dev', 0):.6f}")
                    ws_agg.cell(row=row, column=4, value=f"{agg_stats[metric].get('range', 0):.6f}")
                    row += 1
            
            # Test results
            row += 1
            if "test_results" in analysis:
                ws_agg.cell(row=row, column=1, value="Test Results").font = Font(bold=True, size=11)
                row += 1
                test_results = analysis["test_results"]
                write_metric(ws_agg, row, "KS Uniformity Passed:", test_results.get("ks_uniformity_passed", "N/A"))
                row += 1
                write_metric(ws_agg, row, "NIST Runs Test Passed:", test_results.get("runs_test_passed", "N/A"))
                row += 1
                write_metric(ws_agg, row, "NIST Matrix Rank Passed:", test_results.get("binary_matrix_rank_test_passed", "N/A"))
                row += 1
                write_metric(ws_agg, row, "NIST Longest Run Passed:", test_results.get("longest_run_of_ones_test_passed", "N/A"))
                row += 1
                write_metric(ws_agg, row, "NIST Approx Entropy Passed:", test_results.get("approximate_entropy_test_passed", "N/A"))
                row += 2
            
            # Autocorrelation table
            if "autocorrelation_table" in analysis:
                ws_agg.cell(row=row, column=1, value="Autocorrelation Analysis by Run").font = Font(bold=True, size=11)
                row += 1
                headers = ["Run", "Significant Lags", "Max |Correlation|"]
                for col, header in enumerate(headers, 1):
                    create_header(ws_agg, row, col, header, "B1A0C7")
                row += 1
                for ac_row in analysis["autocorrelation_table"]:
                    ws_agg.cell(row=row, column=1, value=ac_row.get("run", ""))
                    sig_lags = ac_row.get("significant_lags", [])
                    if sig_lags and sig_lags[0] != "None":
                        ws_agg.cell(row=row, column=2, value=", ".join(map(str, sig_lags)))
                    else:
                        ws_agg.cell(row=row, column=2, value="None")
                    ws_agg.cell(row=row, column=3, value=f"{ac_row.get('max_correlation', 0):.6f}")
                    row += 1
                row += 2
            
            # Overlaid ECDF Plot
            if "ecdf_all_runs" in analysis:
                ws_agg.cell(row=row, column=1, value="Overlaid ECDF Plot (All Runs)").font = Font(bold=True, size=11)
                row += 1
                ws_agg.cell(row=row, column=1, value="X").font = Font(bold=True)
                ecdf_col = 2
                ecdf_start_row = row + 1
                # Create unified dataset
                all_x_values = []
                for run_data in analysis["ecdf_all_runs"]:
                    all_x_values.extend(run_data.get("x", []))
                unique_x = sorted(set(all_x_values))
                # Limit to reasonable size
                step = max(1, len(unique_x) // 500) if len(unique_x) > 500 else 1
                unique_x = unique_x[::step]
                
                # Write header row
                for run_data in analysis["ecdf_all_runs"]:
                    run_num = run_data.get("run", ecdf_col - 1)
                    ws_agg.cell(row=row, column=ecdf_col, value=f"Run {run_num}").font = Font(bold=True)
                    ecdf_col += 1
                row += 1
                
                # Write data
                for x_val in unique_x:
                    ws_agg.cell(row=row, column=1, value=f"{x_val:.6f}")
                    col = 2
                    for run_data in analysis["ecdf_all_runs"]:
                        run_x = run_data.get("x", [])
                        run_y = run_data.get("y", [])
                        # Find closest x value and interpolate
                        if run_x:
                            closest_idx = min(range(len(run_x)), key=lambda i: abs(run_x[i] - x_val))
                            if abs(run_x[closest_idx] - x_val) < (max(run_x) - min(run_x)) / len(run_x) * 2:
                                ws_agg.cell(row=row, column=col, value=f"{run_y[closest_idx]:.6f}")
                            else:
                                ws_agg.cell(row=row, column=col, value="")
                        col += 1
                    row += 1
                
                # Add Overlaid ECDF chart
                chart_ecdf_overlay = LineChart()
                chart_ecdf_overlay.title = "Overlaid ECDF Plot (All Runs)"
                chart_ecdf_overlay.y_axis.title = "Cumulative Probability"
                chart_ecdf_overlay.x_axis.title = "Value"
                cats = Reference(ws_agg, min_col=1, min_row=ecdf_start_row, max_row=row-1)
                chart_ecdf_overlay.set_categories(cats)
                # Clear default series
                chart_ecdf_overlay.series = []
                for col_idx in range(2, ecdf_col):
                    yvalues = Reference(ws_agg, min_col=col_idx, min_row=ecdf_start_row, max_row=row-1)
                    title_cell = ws_agg.cell(row=ecdf_start_row-1, column=col_idx)
                    series_title = title_cell.value if title_cell.value else f"Run {col_idx-1}"
                    series = Series(yvalues, title=series_title)
                    chart_ecdf_overlay.series.append(series)
                chart_ecdf_overlay.width = 15
                chart_ecdf_overlay.height = 10
                ws_agg.add_chart(chart_ecdf_overlay, f"F{ecdf_start_row}")
                row += 15
            
            # Overlaid Q-Q Plot
            if "individual_analyses" in analysis:
                row += 2
                ws_agg.cell(row=row, column=1, value="Overlaid Q-Q Plot (All Runs)").font = Font(bold=True, size=11)
                row += 1
                ws_agg.cell(row=row, column=1, value="Theoretical").font = Font(bold=True)
                qq_col = 2
                qq_start_row = row + 1
                # Get all theoretical values
                all_theoretical = set()
                for run_analysis in analysis["individual_analyses"]:
                    if "distribution" in run_analysis and "qq_plot" in run_analysis["distribution"]:
                        all_theoretical.update(run_analysis["distribution"]["qq_plot"].get("theoretical", []))
                unique_theoretical = sorted(all_theoretical)
                step = max(1, len(unique_theoretical) // 500) if len(unique_theoretical) > 500 else 1
                unique_theoretical = unique_theoretical[::step]
                
                # Write header row
                for idx, run_analysis in enumerate(analysis["individual_analyses"]):
                    ws_agg.cell(row=row, column=qq_col, value=f"Run {idx + 1}").font = Font(bold=True)
                    qq_col += 1
                row += 1
                
                # Write data
                for th_val in unique_theoretical:
                    ws_agg.cell(row=row, column=1, value=f"{th_val:.6f}")
                    col = 2
                    for run_analysis in analysis["individual_analyses"]:
                        if "distribution" in run_analysis and "qq_plot" in run_analysis["distribution"]:
                            qq_theoretical = run_analysis["distribution"]["qq_plot"].get("theoretical", [])
                            qq_sample = run_analysis["distribution"]["qq_plot"].get("sample", [])
                            if qq_theoretical:
                                closest_idx = min(range(len(qq_theoretical)), key=lambda i: abs(qq_theoretical[i] - th_val))
                                ws_agg.cell(row=row, column=col, value=f"{qq_sample[closest_idx]:.6f}")
                            else:
                                ws_agg.cell(row=row, column=col, value="")
                        col += 1
                    row += 1
                
                # Add Overlaid Q-Q chart
                chart_qq_overlay = ScatterChart()
                chart_qq_overlay.title = "Overlaid Q-Q Plot (All Runs)"
                chart_qq_overlay.y_axis.title = "Sample"
                chart_qq_overlay.x_axis.title = "Theoretical"
                xvalues = Reference(ws_agg, min_col=1, min_row=qq_start_row, max_row=row-1)
                chart_qq_overlay.style = 13
                # Add series for each run
                for col_idx in range(2, qq_col):
                    yvalues = Reference(ws_agg, min_col=col_idx, min_row=qq_start_row-1, max_row=row-1)
                    series = Series(yvalues, xvalues, title=ws_agg.cell(row=qq_start_row-1, column=col_idx).value)
                    chart_qq_overlay.series.append(series)
                chart_qq_overlay.width = 15
                chart_qq_overlay.height = 10
                ws_agg.add_chart(chart_qq_overlay, f"F{qq_start_row}")
        
        # Sheet 7: Independence Analysis
        if "independence" in analysis:
            ws_ind = wb.create_sheet("Independence", 6)
            row = 1
            create_header(ws_ind, row, 1, "Independence & Correlation Analysis", "C55AA0")
            ws_ind.merge_cells(f"A{row}:B{row}")
            row += 2
            
            ind = analysis["independence"]
            
            # Time series data and chart
            if "time_series" in ind:
                ws_ind.cell(row=row, column=1, value="Time Series Data").font = Font(bold=True, size=11)
                row += 1
                ws_ind.cell(row=row, column=1, value="Index").font = Font(bold=True)
                ws_ind.cell(row=row, column=2, value="Value").font = Font(bold=True)
                row += 1
                indices = ind["time_series"].get("index", [])
                values = ind["time_series"].get("values", [])
                ts_start_row = row
                # Limit to reasonable number of points
                step = max(1, len(indices) // 1000) if len(indices) > 1000 else 1
                for i in range(0, len(indices), step):
                    ws_ind.cell(row=row, column=1, value=indices[i])
                    ws_ind.cell(row=row, column=2, value=f"{values[i]:.6f}")
                    row += 1
                
                # Add Time Series chart
                chart_ts = LineChart()
                chart_ts.title = "Time Series"
                chart_ts.y_axis.title = "Value"
                chart_ts.x_axis.title = "Index"
                data = Reference(ws_ind, min_col=2, min_row=ts_start_row-1, max_row=row-1)
                cats = Reference(ws_ind, min_col=1, min_row=ts_start_row, max_row=row-1)
                chart_ts.add_data(data, titles_from_data=True)
                chart_ts.set_categories(cats)
                chart_ts.width = 15
                chart_ts.height = 10
                ws_ind.add_chart(chart_ts, f"D{ts_start_row}")
                row += 15
            
            # Autocorrelation data and chart
            if "autocorrelation" in ind:
                row += 2
                ws_ind.cell(row=row, column=1, value="Autocorrelation Function (ACF)").font = Font(bold=True, size=11)
                row += 1
                ws_ind.cell(row=row, column=1, value="Lag").font = Font(bold=True)
                ws_ind.cell(row=row, column=2, value="Correlation").font = Font(bold=True)
                row += 1
                lags = ind["autocorrelation"].get("lags", [])
                values = ind["autocorrelation"].get("values", [])
                acf_start_row = row
                for lag, corr in zip(lags, values):
                    ws_ind.cell(row=row, column=1, value=lag)
                    ws_ind.cell(row=row, column=2, value=f"{corr:.6f}")
                    row += 1
                
                # Add ACF chart
                chart_acf = BarChart()
                chart_acf.type = "col"
                chart_acf.style = 10
                chart_acf.title = "Autocorrelation Function (ACF)"
                chart_acf.y_axis.title = "Correlation"
                chart_acf.x_axis.title = "Lag"
                data = Reference(ws_ind, min_col=2, min_row=acf_start_row-1, max_row=row-1)
                cats = Reference(ws_ind, min_col=1, min_row=acf_start_row, max_row=row-1)
                chart_acf.add_data(data, titles_from_data=True)
                chart_acf.set_categories(cats)
                chart_acf.width = 15
                chart_acf.height = 10
                ws_ind.add_chart(chart_acf, f"D{acf_start_row}")
                row += 15
            
            # Lag-1 Scatter plot
            if "lag1_scatter" in ind:
                row += 2
                ws_ind.cell(row=row, column=1, value="Lag-1 Scatter Plot").font = Font(bold=True, size=11)
                row += 1
                ws_ind.cell(row=row, column=1, value="X_n").font = Font(bold=True)
                ws_ind.cell(row=row, column=2, value="X_{n+1}").font = Font(bold=True)
                row += 1
                lag1_x = ind["lag1_scatter"].get("x", [])
                lag1_y = ind["lag1_scatter"].get("y", [])
                lag1_start_row = row
                # Limit to reasonable number of points
                step = max(1, len(lag1_x) // 1000) if len(lag1_x) > 1000 else 1
                for i in range(0, len(lag1_x), step):
                    ws_ind.cell(row=row, column=1, value=f"{lag1_x[i]:.6f}")
                    ws_ind.cell(row=row, column=2, value=f"{lag1_y[i]:.6f}")
                    row += 1
                
                # Add Lag-1 Scatter chart
                chart_lag1 = ScatterChart()
                chart_lag1.title = "Lag-1 Scatter Plot"
                chart_lag1.y_axis.title = "X_{n+1}"
                chart_lag1.x_axis.title = "X_n"
                chart_lag1.style = 13
                xvalues = Reference(ws_ind, min_col=1, min_row=lag1_start_row, max_row=row-1)
                yvalues = Reference(ws_ind, min_col=2, min_row=lag1_start_row-1, max_row=row-1)
                series = Series(yvalues, xvalues, title="Lag-1")
                chart_lag1.series.append(series)
                chart_lag1.width = 15
                chart_lag1.height = 10
                ws_ind.add_chart(chart_lag1, f"D{lag1_start_row}")
        
        # Sheet 8: Stationarity Analysis
        if "stationarity" in analysis:
            ws_stat = wb.create_sheet("Stationarity", 7)
            row = 1
            create_header(ws_stat, row, 1, "Stationarity Analysis", "00B0F0")
            ws_stat.merge_cells(f"A{row}:B{row}")
            row += 2
            
            stat = analysis["stationarity"]
            
            # Rolling mean and std data and chart
            if "rolling_mean" in stat and "rolling_std" in stat:
                ws_stat.cell(row=row, column=1, value="Rolling Mean & Standard Deviation").font = Font(bold=True, size=11)
                row += 1
                ws_stat.cell(row=row, column=1, value="Index").font = Font(bold=True)
                ws_stat.cell(row=row, column=2, value="Rolling Mean").font = Font(bold=True)
                ws_stat.cell(row=row, column=3, value="Rolling Std").font = Font(bold=True)
                row += 1
                rolling_idx = stat["rolling_mean"].get("index", [])
                rolling_mean = stat["rolling_mean"].get("values", [])
                rolling_std = stat["rolling_std"].get("values", [])
                rolling_start_row = row
                # Limit to reasonable number of points
                step = max(1, len(rolling_idx) // 1000) if len(rolling_idx) > 1000 else 1
                for i in range(0, len(rolling_idx), step):
                    ws_stat.cell(row=row, column=1, value=rolling_idx[i])
                    ws_stat.cell(row=row, column=2, value=f"{rolling_mean[i]:.6f}")
                    ws_stat.cell(row=row, column=3, value=f"{rolling_std[i]:.6f}")
                    row += 1
                
                # Add Rolling Stats chart
                chart_rolling = LineChart()
                chart_rolling.title = "Rolling Mean & Standard Deviation"
                chart_rolling.y_axis.title = "Value"
                chart_rolling.x_axis.title = "Index"
                cats = Reference(ws_stat, min_col=1, min_row=rolling_start_row, max_row=row-1)
                chart_rolling.set_categories(cats)
                data_mean = Reference(ws_stat, min_col=2, min_row=rolling_start_row-1, max_row=row-1)
                series_mean = chart_rolling.series.append(data_mean)
                series_mean.title = "Rolling Mean"
                data_std = Reference(ws_stat, min_col=3, min_row=rolling_start_row-1, max_row=row-1)
                series_std = chart_rolling.series.append(data_std)
                series_std.title = "Rolling Std"
                chart_rolling.width = 15
                chart_rolling.height = 10
                ws_stat.add_chart(chart_rolling, f"E{rolling_start_row}")
                row += 15
            
            # Chunked statistics
            if "chunks" in stat:
                row += 2
                ws_stat.cell(row=row, column=1, value="Chunked Statistics").font = Font(bold=True, size=11)
                row += 1
                headers = ["Chunk", "Mean", "Std Dev", "Min", "Max"]
                for col, header in enumerate(headers, 1):
                    create_header(ws_stat, row, col, header, "5B9BD5")
                row += 1
                for chunk in stat["chunks"]:
                    ws_stat.cell(row=row, column=1, value=chunk.get("chunk"))
                    ws_stat.cell(row=row, column=2, value=f"{chunk.get('mean', 0):.6f}")
                    ws_stat.cell(row=row, column=3, value=f"{chunk.get('std', 0):.6f}")
                    ws_stat.cell(row=row, column=4, value=f"{chunk.get('min', 0):.6f}")
                    ws_stat.cell(row=row, column=5, value=f"{chunk.get('max', 0):.6f}")
                    row += 1
        
        # Sheet 9: Spectral Analysis
        if "spectral" in analysis:
            ws_spec = wb.create_sheet("Spectral", 8)
            row = 1
            create_header(ws_spec, row, 1, "Spectral Analysis", "FF6600")
            ws_spec.merge_cells(f"A{row}:D{row}")
            row += 2
            
            spec = analysis["spectral"]
            headers = ["Frequency", "Magnitude", "Power"]
            for col, header in enumerate(headers, 1):
                create_header(ws_spec, row, col, header, "FF9900")
            row += 1
            
            freqs = spec.get("frequencies", [])
            magnitudes = spec.get("magnitude", [])
            powers = spec.get("power", [])
            spec_start_row = row
            # Limit to reasonable number of points
            step = max(1, len(freqs) // 1000) if len(freqs) > 1000 else 1
            for i in range(0, len(freqs), step):
                ws_spec.cell(row=row, column=1, value=f"{freqs[i]:.6f}")
                ws_spec.cell(row=row, column=2, value=f"{magnitudes[i]:.6f}")
                ws_spec.cell(row=row, column=3, value=f"{powers[i]:.6f}")
                row += 1
            
            # Add FFT Magnitude chart
            chart_mag = LineChart()
            chart_mag.title = "FFT Magnitude"
            chart_mag.y_axis.title = "Magnitude"
            chart_mag.x_axis.title = "Frequency"
            data = Reference(ws_spec, min_col=2, min_row=spec_start_row-1, max_row=row-1)
            cats = Reference(ws_spec, min_col=1, min_row=spec_start_row, max_row=row-1)
            chart_mag.add_data(data, titles_from_data=True)
            chart_mag.set_categories(cats)
            chart_mag.width = 15
            chart_mag.height = 10
            ws_spec.add_chart(chart_mag, f"E{spec_start_row}")
            
            # Add Power Spectrum chart
            chart_power = LineChart()
            chart_power.title = "Power Spectrum (Periodogram)"
            chart_power.y_axis.title = "Power"
            chart_power.x_axis.title = "Frequency"
            data = Reference(ws_spec, min_col=3, min_row=spec_start_row-1, max_row=row-1)
            chart_power.add_data(data, titles_from_data=True)
            chart_power.set_categories(cats)
            chart_power.width = 15
            chart_power.height = 10
            ws_spec.add_chart(chart_power, f"E{spec_start_row + 15}")
        
        # Sheet 10: NIST Tests
        if "nist_tests" in analysis:
            ws_nist = wb.create_sheet("NIST Tests", 9)
            row = 1
            create_header(ws_nist, row, 1, "NIST Statistical Tests", "E74C3C")
            ws_nist.merge_cells(f"A{row}:B{row}")
            row += 2
            
            nist = analysis["nist_tests"]
            
            # Runs Test
            if "runs_test" in nist:
                ws_nist.cell(row=row, column=1, value="Runs Test").font = Font(bold=True, size=11)
                row += 1
                runs_test = nist["runs_test"]
                if "error" in runs_test and runs_test["error"]:
                    write_metric(ws_nist, row, "Error:", runs_test["error"])
                    row += 1
                else:
                    write_metric(ws_nist, row, "P-value:", f"{runs_test.get('p_value', 0):.6f}")
                    row += 1
                    write_metric(ws_nist, row, "Z-statistic:", f"{runs_test.get('statistic', 0):.4f}")
                    row += 1
                    write_metric(ws_nist, row, "Runs observed:", runs_test.get('runs', 'N/A'))
                    row += 1
                    write_metric(ws_nist, row, "Runs expected:", f"{runs_test.get('expected_runs', 0):.4f}")
                    row += 1
                    write_metric(ws_nist, row, "Ones:", runs_test.get('ones', 'N/A'))
                    row += 1
                    write_metric(ws_nist, row, "Zeros:", runs_test.get('zeros', 'N/A'))
                    row += 1
                    write_metric(ws_nist, row, "Result:", "Passed" if runs_test.get('passed', False) else "Failed")
                    row += 1
                row += 1
            
            # Binary Matrix Rank Test
            if "binary_matrix_rank_test" in nist:
                ws_nist.cell(row=row, column=1, value="Binary Matrix Rank Test").font = Font(bold=True, size=11)
                row += 1
                matrix_test = nist["binary_matrix_rank_test"]
                if "error" in matrix_test and matrix_test["error"]:
                    write_metric(ws_nist, row, "Error:", matrix_test["error"])
                    row += 1
                else:
                    write_metric(ws_nist, row, "P-value:", f"{matrix_test.get('p_value', 0):.6f}")
                    row += 1
                    write_metric(ws_nist, row, "Chi-square statistic:", f"{matrix_test.get('statistic', 0):.4f}")
                    row += 1
                    write_metric(ws_nist, row, "Number of matrices:", matrix_test.get('num_matrices', 'N/A'))
                    row += 1
                    write_metric(ws_nist, row, "Full rank count:", matrix_test.get('full_rank_count', 'N/A'))
                    row += 1
                    write_metric(ws_nist, row, "Rank-1 count:", matrix_test.get('rank_minus_1_count', 'N/A'))
                    row += 1
                    write_metric(ws_nist, row, "Rank-0 count:", matrix_test.get('rank_0_count', 'N/A'))
                    row += 1
                    write_metric(ws_nist, row, "Result:", "Passed" if matrix_test.get('passed', False) else "Failed")
                    row += 1
                row += 1
            
            # Longest Run of Ones Test
            if "longest_run_of_ones_test" in nist:
                ws_nist.cell(row=row, column=1, value="Longest Run of Ones Test").font = Font(bold=True, size=11)
                row += 1
                longest_run_test = nist["longest_run_of_ones_test"]
                if "error" in longest_run_test and longest_run_test["error"]:
                    write_metric(ws_nist, row, "Error:", longest_run_test["error"])
                    row += 1
                else:
                    write_metric(ws_nist, row, "P-value:", f"{longest_run_test.get('p_value', 0):.6f}")
                    row += 1
                    write_metric(ws_nist, row, "Chi-square statistic:", f"{longest_run_test.get('statistic', 0):.4f}")
                    row += 1
                    write_metric(ws_nist, row, "Number of blocks:", longest_run_test.get('num_blocks', 'N/A'))
                    row += 1
                    if "run_counts" in longest_run_test:
                        ws_nist.cell(row=row, column=1, value="Run length counts:").font = Font(bold=True)
                        row += 1
                        for length, count in longest_run_test["run_counts"].items():
                            write_metric(ws_nist, row, f"Length ≤{length}:", count)
                            row += 1
                    write_metric(ws_nist, row, "Result:", "Passed" if longest_run_test.get('passed', False) else "Failed")
                    row += 1
                row += 1
            
            # Approximate Entropy Test
            if "approximate_entropy_test" in nist:
                ws_nist.cell(row=row, column=1, value="Approximate Entropy Test").font = Font(bold=True, size=11)
                row += 1
                apen_test = nist["approximate_entropy_test"]
                if "error" in apen_test and apen_test["error"]:
                    write_metric(ws_nist, row, "Error:", apen_test["error"])
                    row += 1
                else:
                    write_metric(ws_nist, row, "P-value:", f"{apen_test.get('p_value', 0):.6f}")
                    row += 1
                    write_metric(ws_nist, row, "Chi-square statistic:", f"{apen_test.get('statistic', 0):.4f}")
                    row += 1
                    write_metric(ws_nist, row, "Approximate Entropy:", f"{apen_test.get('approximate_entropy', 0):.6f}")
                    row += 1
                    write_metric(ws_nist, row, "Phi(m):", f"{apen_test.get('phi_m', 0):.6f}")
                    row += 1
                    write_metric(ws_nist, row, "Phi(m+1):", f"{apen_test.get('phi_m1', 0):.6f}")
                    row += 1
                    write_metric(ws_nist, row, "Pattern length m:", apen_test.get('pattern_length_m', 'N/A'))
                    row += 1
                    write_metric(ws_nist, row, "Pattern length m+1:", apen_test.get('pattern_length_m1', 'N/A'))
                    row += 1
                    write_metric(ws_nist, row, "Unique patterns (m):", apen_test.get('unique_patterns_m', 'N/A'))
                    row += 1
                    write_metric(ws_nist, row, "Unique patterns (m+1):", apen_test.get('unique_patterns_m1', 'N/A'))
                    row += 1
                    write_metric(ws_nist, row, "Result:", "Passed" if apen_test.get('passed', False) else "Failed")
                    row += 1
                row += 1
            
            # Binary Sequence Information
            if "binary_sequence_length" in nist:
                row += 1
                ws_nist.cell(row=row, column=1, value="Binary Sequence Information").font = Font(bold=True, size=11)
                row += 1
                write_metric(ws_nist, row, "Total binary sequence length:", f"{nist['binary_sequence_length']:,} bits")
                row += 1
        
        # Sheet 11: Per-Run Analysis (if multi-run)
        if "individual_analyses" in analysis:
            ws_per_run = wb.create_sheet("Per-Run Analysis", 10)
            row = 1
            create_header(ws_per_run, row, 1, "Per-Run Statistics Summary", "0070C0")
            ws_per_run.merge_cells(f"A{row}:G{row}")
            row += 2
            
            headers = ["Run", "Mean", "Std Dev", "Min", "Max", "Range", "KS Test (p)"]
            for col, header in enumerate(headers, 1):
                create_header(ws_per_run, row, col, header, "4F81BD")
            row += 1
            
            for idx, run_analysis in enumerate(analysis["individual_analyses"]):
                basic_stats = run_analysis.get("basic_stats", {})
                dist = run_analysis.get("distribution", {})
                ws_per_run.cell(row=row, column=1, value=f"Run {idx + 1}")
                ws_per_run.cell(row=row, column=2, value=f"{basic_stats.get('mean', 0):.6f}")
                ws_per_run.cell(row=row, column=3, value=f"{basic_stats.get('std', 0):.6f}")
                ws_per_run.cell(row=row, column=4, value=f"{basic_stats.get('min', 0):.6f}")
                ws_per_run.cell(row=row, column=5, value=f"{basic_stats.get('max', 0):.6f}")
                range_val = basic_stats.get('max', 0) - basic_stats.get('min', 0)
                ws_per_run.cell(row=row, column=6, value=f"{range_val:.6f}")
                ks_p = dist.get("is_uniform", {}).get("ks_p", "N/A")
                ws_per_run.cell(row=row, column=7, value=f"{ks_p:.6f}" if isinstance(ks_p, (int, float)) else "N/A")
                row += 1
        
        # Auto-adjust column widths
        for ws in wb.worksheets:
            for col_idx, column in enumerate(ws.columns, start=1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                
                # Calculate max length, skipping MergedCell objects
                for cell in column:
                    try:
                        # MergedCell objects don't have a 'value' attribute
                        # Check if it's a regular cell by trying to access value
                        if hasattr(cell, 'value'):
                            cell_value = cell.value
                            if cell_value is not None:
                                cell_str = str(cell_value)
                                if len(cell_str) > max_length:
                                    max_length = len(cell_str)
                    except (AttributeError, TypeError):
                        # Skip MergedCell objects and other non-standard cells
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"analysis_report_{analysis.get('provider', 'unknown')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Invalid JSON in parameters")
    except Exception as e:
        logger.error(f"Error generating XLSX: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error generating XLSX: {str(e)}")


class PDFDownloadRequest(BaseModel):
    analysis: Dict[str, Any]
    runs: List[List[float]]

@app.post("/download/pdf")
async def download_pdf(request: PDFDownloadRequest):
    """Download full analysis report as PDF file with metrics, charts, and everything using LaTeX"""
    temp_dir = None
    try:
        analysis = request.analysis
        runs_data = request.runs
        
        # Create temporary directory for LaTeX compilation
        temp_dir = tempfile.mkdtemp()
        
        # Generate PDF using LaTeX
        pdf_bytes = generate_latex_pdf(analysis, runs_data, temp_dir)
        
        filename = f"LLM_Randomness_Evaluation_Report_{analysis.get('provider', 'unknown')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        logger.error(f"Error generating PDF: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error generating PDF: {str(e)}")
    finally:
        # Clean up temporary directory
        if temp_dir and os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)


@app.get("/dummy-data")
async def get_dummy_data():
    """Get dummy data from file for testing"""
    try:
        dummy_file_path = os.path.join(os.path.dirname(__file__), "data", DUMMY_DATA_FILENAME)
        if not os.path.exists(dummy_file_path):
            raise HTTPException(status_code=404, detail="Dummy data file not found")
        
        with open(dummy_file_path, 'r') as f:
            data = json.load(f)
        
        # Determine if it's single run or multi-run
        if isinstance(data, list) and len(data) > 0:
            if isinstance(data[0], list):
                # Multi-run format: [[x,x,x], [x,x,x]]
                return {
                    "data": data,
                    "is_multi_run": True,
                    "num_runs": len(data),
                    "count_per_run": len(data[0]) if data else 0
                }
            else:
                # Single run format: [x,x,x]
                return {
                    "data": [data],  # Wrap in array for consistency
                    "is_multi_run": False,
                    "num_runs": 1,
                    "count_per_run": len(data)
                }
        else:
            raise HTTPException(status_code=400, detail="Invalid dummy data format")
    except json.JSONDecodeError as e:
        logger.error(f"Error parsing dummy data file: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error parsing dummy data: {str(e)}")
    except Exception as e:
        logger.error(f"Error loading dummy data: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error loading dummy data: {str(e)}")


@app.get("/dummy-data/stream")
async def stream_dummy_data():
    """Stream dummy data as SSE events to match LLM streaming format"""
    dummy_file_path = os.path.join(os.path.dirname(__file__), "data", DUMMY_DATA_FILENAME)
    
    if not os.path.exists(dummy_file_path):
        async def error_generator():
            yield f"data: {json.dumps({'error': 'Dummy data file not found'})}\n\n"
            yield "data: [DONE]\n\n"
        return StreamingResponse(
            error_generator(),
            media_type="text/event-stream",
            headers={
                "Cache-Control": "no-cache",
                "Connection": "keep-alive",
                "X-Accel-Buffering": "no",
            }
        )
    
    async def event_generator():
        try:
            with open(dummy_file_path, 'r') as f:
                data = json.load(f)
            
            # Determine format
            if isinstance(data, list) and len(data) > 0:
                if isinstance(data[0], list):
                    # Multi-run: stream all runs, with [DONE] between runs
                    for run_idx, run in enumerate(data):
                        for number in run:
                            yield f"data: {json.dumps({'number': number, 'provider': 'dummy'})}\n\n"
                            await asyncio.sleep(0.01)  # Small delay to simulate streaming
                        # Send [DONE] after each run (except the last one)
                        if run_idx < len(data) - 1:
                            yield "data: [DONE]\n\n"
                else:
                    # Single run: stream the array
                    for number in data:
                        yield f"data: {json.dumps({'number': number, 'provider': 'dummy'})}\n\n"
                        await asyncio.sleep(0.01)  # Small delay to simulate streaming
            # Final [DONE] to signal end of all data
            yield "data: [DONE]\n\n"
        except Exception as e:
            logger.error(f"Error in dummy data stream: {str(e)}", exc_info=True)
            yield f"data: {json.dumps({'error': str(e)})}\n\n"
            yield "data: [DONE]\n\n"
    
    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        }
    )


@app.get("/providers")
async def get_providers():
    """Get list of available LLM providers"""
    return {
        "providers": [
            {"id": "openai", "name": "OpenAI"},
            {"id": "anthropic", "name": "Anthropic"},
            {"id": "deepseek", "name": "DeepSeek"}
        ]
    }


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
